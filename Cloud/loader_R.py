import mysql.connector
from openpyxl import Workbook, load_workbook


db = mysql.connector.connect(
    host='34.138.188.252',
    user='root',
    passwd='uconnstamford',
    database='DATA'
    )

mycursor = db.cursor()

##create the database
##this should only be executed before the creation of the DATA database

#mycursor.execute("CREATE DATABASE DATA")


##create table to input merged_data.csv
##this should only be executed once
#str_ = 'CREATE TABLE MSCI_R (isinID VARCHAR(40) NOT NULL, ticker VARCHAR(50) DEFAULT NULL, Month smallint CHECK(Month BETWEEN 1 AND 12), Year int, price float(8,2), environmentalPillarScore float(5,2), goverancePillarScore float(5,2), socialPillarScore float(5,2))'

#mycursor.execute(str_)


##Insert MSCI_Price_Data.xlsx rows into database

wb = load_workbook('MSCI_Price_Data.xlsx')
ws = wb.active


for i in range(2, ws.max_row + 1):

    mycursor.execute("INSERT INTO MSCI_R (isinID, ticker, Month, Year, price, environmentalPillarScore, goverancePillarScore, socialPillarScore) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)", (ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value, ws['E' + str(i)].value, ws['F' + str(i)].value, ws['G' + str(i)].value, ws['H' + str(i)].value, ws['I' + str(i)].value))
    db.commit()

    print("Loading record {}".format(i))
