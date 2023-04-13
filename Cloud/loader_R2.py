import mysql.connector
from openpyxl import Workbook, load_workbook


db = mysql.connector.connect(
    host='34.138.188.252',
    user='root',
    passwd='uconnstamford',
    database='DATA'
    )

mycursor = db.cursor()

##create the database
##this should only be executed before the creation of the DATA database

#mycursor.execute("CREATE DATABASE DATA")


##create table to input merged_data.csv
##this should only be executed once
#str_ = 'CREATE TABLE PRICE_R (DATE_DIM_ID VARCHAR(40) NOT NULL, BB_TICKER_CD VARCHAR(50) DEFAULT NULL, ISIN_CD VARCHAR(50), PBD_PRICE_AMT float(8,2))'

#mycursor.execute(str_)


##Insert PriceData_2012_2022.xlsx rows into database

wb = load_workbook('PriceData_2012_2022.xlsx')
ws = wb.active


for i in range(30000, ws.max_row + 1):

    price = str(ws['D' + str(i)].value)
    ind = price.find('.')

    if ind == -1: price += ".00"
    elif ((len(price) - 1) - ind) < 2: price += "0"

    try: price = float(price)
    except: continue
    
    mycursor.execute("INSERT INTO PRICE_R (DATE_DIM_ID, BB_TICKER_CD, ISIN_CD, PBD_PRICE_AMT) VALUES (%s, %s, %s, %s)", (ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value, price))
    db.commit()

    print("Loading record {}".format(i))
