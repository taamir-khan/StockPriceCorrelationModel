import mysql.connector
from openpyxl import Workbook, load_workbook


db = mysql.connector.connect(
    host='34.138.188.252',
    user='root',
    passwd='uconnstamford',
    database='R_DATA'
    )

mycursor = db.cursor()

##create the database
##this should only be executed before the creation of the R_DATA database

#mycursor.execute("CREATE DATABASE R_DATA")


##create table to input merged_data.csv
##this should only be executed once
#str_ = 'CREATE TABLE MSCI_R (ASOF_DATE VARCHAR(30), ISSUER_ISIN VARCHAR(100), INDUSTRY_ADJUSTED_SCORE FLOAT(8,2), WEIGHTED_AVERAGE_SCORE FLOAT(8,2), IVA_INDUSTRY VARCHAR(70), ENVIRONMENTAL_PILLAR_SCORE FLOAT(8,2), GOVERNANCE_PILLAR_SCORE FLOAT(8,2), SOCIAL_PILLAR_SCORE FLOAT(8,2))'

#mycursor.execute(str_)


##Insert MSCIData_2014_to_2018_2021_to_2022.xlsx rows into database

wb = load_workbook('MSCIData_2014_to_2018_2021_to_2022.xlsx')
ws = wb.active


for i in range(2, ws.max_row + 1):

#    price = str(ws['D' + str(i)].value)
#    ind = price.find('.')

#    if ind == -1: price += ".00"
#    elif ((len(price) - 1) - ind) < 2: price += "0"

#    try: price = float(price)
#    except: continue
    
    mycursor.execute("INSERT INTO MSCI_R (ASOF_DATE, ISSUER_ISIN, INDUSTRY_ADJUSTED_SCORE, WEIGHTED_AVERAGE_SCORE, IVA_INDUSTRY, ENVIRONMENTAL_PILLAR_SCORE, GOVERNANCE_PILLAR_SCORE, SOCIAL_PILLAR_SCORE) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)", (ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value, ws['E' + str(i)].value, ws['F' + str(i)].value, ws['G' + str(i)].value, ws['H' + str(i)].value))
    db.commit()

    print("Loading record {}".format(i))
