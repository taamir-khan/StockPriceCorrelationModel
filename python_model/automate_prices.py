from openpyxl import Workbook
from openpyxl import load_workbook


loaded_book = 'PricesAES.xlsx'
row_ = 1486
new_book = 'testAES.xlsx'

###find all averages based on month/year
to_month = {'01': 'January', '02': 'February', '03': 'March',
            '04': 'April', '05': 'May', '06': 'June',
            '07': 'July', '08': 'August', '09': 'September',
            '10': 'October', '11': 'November', '12': 'December'}

wb = load_workbook(loaded_book)
ws = wb.active

curr_month = curr_year = None
all_monthNyears = []
sum_ = count_ = 0
for i in range(row_, 1, -1):

    temp = str(ws['A' + str(i)].value)
    price = str(ws['B' + str(i)].value)
    year, month = temp[0:4], to_month[temp[5:7]]

    if curr_month == curr_year == None:
        
        sum_ = float(price)
        count_ = 1
        curr_month = month
        curr_year = year
    
    elif (year != curr_year) or (month != curr_month):
    
        all_monthNyears.append((curr_month, curr_year, sum_ / float(count_)))
        sum_ = float(price)
        count_ = 1
        curr_month = month
        curr_year = year

    else:
    
        sum_ += float(price)
        count_ += 1


all_monthNyears.append((curr_month, curr_year, sum_ / float(count_)))


#### create new workbook
new_wb = Workbook()

new_ws = new_wb.active
new_ws.title = 'Prices'
new_ws['A1'] = 'Month, Year'
new_ws['B1'] = 'Average Price (in dollars)'

count = 2
for pair in all_monthNyears:

    month, year, price = pair
    
    new_ws['A' + str(count)] = '{}, {}'.format(month, year)
    new_ws['B' + str(count)] = '{}'.format(round(price, 2))
    count += 1

new_wb.save(new_book)